import json
import pandas as pd
from openpyxl.styles import Font
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog

# Create a Tkinter window and hide it
root = tk.Tk()
root.withdraw()

# Ask user to select a JSON file
json_file_path = filedialog.askopenfilename(title='Select a JSON file', filetypes=[('JSON files', '*.json')])

# Read JSON data into Python object
with open(json_file_path, 'r', encoding='utf-8') as file:
    data = json.load(file)

# Convert Python object to Pandas DataFrame
df = pd.DataFrame(data)

# Ask user to select a path to save XLSX file
xlsx_file_path = filedialog.asksaveasfilename(title='Save as XLSX', defaultextension='.xlsx', filetypes=[('Excel files', '*.xlsx')])

# Save DataFrame as XLSX file
df.to_excel(xlsx_file_path, index=False)

# Open generated XLSX file
wb = load_workbook(xlsx_file_path)
ws = wb.active

# Set font of entire sheet to bold
font = Font(name='黑体')
for row in ws.iter_rows():
    for cell in row:
        cell.font = font

# Save modified XLSX file
wb.save(xlsx_file_path)
